# Excelkeywords.py

from robot.api.deco import keyword
from robot.libraries.BuiltIn import BuiltIn
from openpyxl import load_workbook

class ExcelKeywords:
    def __init__(self):
        self.builtin = BuiltIn()
    
    @keyword
    def read_excel_cell(self, file_path, sheet_name, row_number, column_number):
        """Reads data from the specified cell in an Excel file."""
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        cell_value = sheet.cell(row=row_number, column=column_number).value
        workbook.close()
        return cell_value
